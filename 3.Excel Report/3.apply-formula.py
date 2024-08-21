from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle

wb = load_workbook('barchart.xlsx')
sheet = wb['Report']

min_row = sheet.min_row
max_row = sheet.max_row
min_column = sheet.min_column
max_column = sheet.max_column

# Define a currency style
currency_style = NamedStyle(name='currency_style', number_format='$#,##0.00')

for i in range(min_column + 1, max_column + 1):
    letter = get_column_letter(i)
    # Create the cell reference for the formula
    cell_ref = f'{letter}{max_row + 1}'
    
    # Apply formula to each cell in the new row
    sheet[cell_ref] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    sheet[cell_ref].style = currency_style

wb.save('report.xlsx')
