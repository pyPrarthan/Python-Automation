from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle

# Define month and read workbook
month = 'february'
wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

# Active rows and columns
min_column = sheet.min_column
max_column = sheet.max_column
min_row = sheet.min_row
max_row = sheet.max_row

# Instantiate a barchart
barchart = BarChart()

# Locate data and categories
data = Reference(sheet, min_col=min_column + 1, max_col=max_column, min_row=min_row, max_row=max_row)  # including headers
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row)  # not including headers

# Adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# Make chart
sheet.add_chart(barchart)
barchart.title = 'Sales by Product line'
barchart.style = 5  # choose the chart style

# Define a currency style
currency_style = NamedStyle(name='currency_style', number_format='$#,##0.00')

# Write multiple formulas and apply style
for i in range(min_column + 1, max_column + 1):
    letter = get_column_letter(i)
    cell_ref = f'{letter}{max_row + 1}'
    sheet[cell_ref] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    sheet[cell_ref].style = currency_style

# Add format
sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font = Font(name='Arial', bold=True, size=20)
sheet['A2'].font = Font(name='Arial', bold=True, size=10)

wb.save(f'report_{month}.xlsx')
