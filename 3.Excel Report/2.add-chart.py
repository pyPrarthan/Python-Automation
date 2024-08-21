from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Load the workbook and select the desired sheet
wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

# Define the boundaries for the data and categories
min_column = sheet.min_column
max_column = sheet.max_column
min_row = sheet.min_row
max_row = sheet.max_row

# Create a BarChart object
barchart = BarChart()

# Define data and categories for the chart
data = Reference(sheet, min_col=min_column + 1, max_col=max_column, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row)

# Add data and categories to the chart
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# Set the title and style for the chart
barchart.title = "Sales by Product line"
barchart.style = 5

# Add the chart to the sheet at the specified location
#sheet.add_chart(barchart,"A1")

# Save the workbook
wb.save('barchart.xlsx')
